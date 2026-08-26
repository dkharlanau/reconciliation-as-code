from __future__ import annotations

import copy
import csv
import hashlib
import html
import json
from pathlib import Path
from typing import Any

from .errors import DataError


def _json_text(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)


def _hash_value(value: Any) -> str:
    digest = hashlib.sha256(_json_text(value).encode("utf-8")).hexdigest()
    return f"sha256:{digest}"


def _protect_value(value: Any, mode: str) -> Any:
    if value is None:
        return None
    if mode == "hash":
        return _hash_value(value)
    if mode == "omit":
        return None
    return "***"


def _sanitize_predicate(predicate: Any, sensitive_fields: set[str], mode: str) -> Any:
    if not isinstance(predicate, dict):
        return predicate
    copied = copy.deepcopy(predicate)
    for logical in ("all", "any"):
        if logical in copied:
            copied[logical] = [
                _sanitize_predicate(item, sensitive_fields, mode) for item in copied[logical]
            ]
            return copied
    if "not" in copied:
        copied["not"] = _sanitize_predicate(copied["not"], sensitive_fields, mode)
        return copied
    if copied.get("field") in sensitive_fields and "value" in copied:
        value = copied["value"]
        if isinstance(value, list):
            copied["value"] = [_protect_value(item, mode) for item in value]
        else:
            copied["value"] = _protect_value(value, mode)
    return copied


def prepare_evidence(result: dict[str, Any], spec: dict[str, Any]) -> dict[str, Any]:
    """Return a presentation-safe copy of canonical evidence according to evidence privacy policy."""
    prepared = copy.deepcopy(result)
    evidence_config = spec.get("evidence") or {}
    sensitive_fields = set(evidence_config.get("sensitive_fields", []))
    sensitive_mode = evidence_config.get("sensitive_value_mode", "mask")
    key_mode = evidence_config.get("key_mode", "plain")

    prepared["privacy"] = {
        "sensitive_fields": sorted(sensitive_fields),
        "sensitive_value_mode": sensitive_mode,
        "key_mode": key_mode,
    }

    for endpoint in ("source", "target"):
        selection = (prepared.get("selection") or {}).get(endpoint)
        if selection and selection.get("filter"):
            selection["filter"] = _sanitize_predicate(
                selection["filter"], sensitive_fields, sensitive_mode
            )

    spec_checks: dict[str, dict[str, Any]] = {}
    for index, check in enumerate(spec.get("checks", []), start=1):
        spec_checks[check.get("id", f"check-{index}")] = check

    for check in prepared.get("checks", []):
        check_spec = spec_checks.get(check["id"], {})
        source_field = check_spec.get("source") or check.get("metrics", {}).get("source_field")
        target_field = check_spec.get("target") or check.get("metrics", {}).get("target_field")
        source_sensitive = source_field in sensitive_fields
        target_sensitive = target_field in sensitive_fields

        if check["type"] == "control_total":
            if source_sensitive:
                check["metrics"]["source_total"] = _protect_value(
                    check["metrics"].get("source_total"), sensitive_mode
                )
            if target_sensitive:
                check["metrics"]["target_total"] = _protect_value(
                    check["metrics"].get("target_total"), sensitive_mode
                )

        for detail in check.get("details", []):
            if key_mode == "hash" and "key" in detail:
                detail["key"] = _hash_value(detail["key"])
            if source_sensitive:
                for name in ("source", "normalized_source", "source_value"):
                    if name in detail:
                        detail[name] = _protect_value(detail[name], sensitive_mode)
            if target_sensitive:
                for name in ("target", "normalized_target", "target_value"):
                    if name in detail:
                        detail[name] = _protect_value(detail[name], sensitive_mode)
            if source_sensitive or target_sensitive:
                detail["sensitive_values_redacted"] = True

    return prepared


def write_json(result: dict[str, Any], path: str | Path) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(result, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def render_markdown(result: dict[str, Any]) -> str:
    summary = result["summary"]
    lines = [
        f"# Reconciliation evidence: {result['reconciliation']}",
        "",
        f"**Status:** `{result['status'].upper()}`  ",
        f"**Run ID:** `{result.get('run', {}).get('id', '')}`  ",
        f"**Generated:** {result['generated_at']}",
        "",
        "## Summary",
        "",
        "| Metric | Value |",
        "| --- | ---: |",
        f"| Source records | {summary['source_records']} |",
        f"| Target records | {summary['target_records']} |",
        f"| Matched records | {summary['matched_records']} |",
        f"| Missing in target | {summary['missing_in_target']} |",
        f"| Unexpected in target | {summary['unexpected_in_target']} |",
        f"| Failed error checks | {summary['checks_failed']} |",
        f"| Failed warning checks | {summary['warnings_failed']} |",
        "",
        "## Checks",
        "",
        "| Check | Type | Severity | Status |",
        "| --- | --- | --- | --- |",
    ]
    for check in result["checks"]:
        lines.append(
            f"| `{check['id']}` | {check['type']} | {check['severity']} | **{check['status']}** |"
        )

    for check in result["checks"]:
        if check["status"] == "passed" and not check["details"]:
            continue
        lines.extend(["", f"### {check['id']}", "", "Metrics:", "", "```json"])
        lines.append(json.dumps(check["metrics"], indent=2, ensure_ascii=False))
        lines.append("```")
        if check["details"]:
            lines.extend(["", "Evidence sample:", "", "```json"])
            lines.append(json.dumps(check["details"], indent=2, ensure_ascii=False))
            lines.append("```")
            if check.get("details_truncated"):
                lines.append("\n_Detail output was truncated by `evidence.detail_limit`._")

    privacy = result.get("privacy") or {}
    if privacy.get("sensitive_fields") or privacy.get("key_mode") == "hash":
        lines.extend(["", "## Privacy controls", ""])
        lines.append(f"- Sensitive value mode: `{privacy.get('sensitive_value_mode', 'mask')}`")
        lines.append(f"- Key mode: `{privacy.get('key_mode', 'plain')}`")
        if privacy.get("sensitive_fields"):
            lines.append("- Protected fields: " + ", ".join(f"`{item}`" for item in privacy["sensitive_fields"]))

    lines.extend(["", "## Input fingerprints", ""])
    for name, info in result.get("inputs", {}).items():
        lines.append(f"- **{name}** — `{info['sha256']}` — `{info['path']}`")
    lines.append("")
    return "\n".join(lines)


def write_markdown(result: dict[str, Any], path: str | Path) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(render_markdown(result), encoding="utf-8")


def _table_html(headers: list[str], rows: list[list[Any]]) -> str:
    head = "".join(f"<th>{html.escape(str(value))}</th>" for value in headers)
    body_rows = []
    for row in rows:
        body_rows.append(
            "<tr>" + "".join(f"<td>{html.escape(str(value if value is not None else ''))}</td>" for value in row) + "</tr>"
        )
    return f"<table><thead><tr>{head}</tr></thead><tbody>{''.join(body_rows)}</tbody></table>"


def render_html(result: dict[str, Any]) -> str:
    summary_rows = [[name.replace("_", " ").title(), value] for name, value in result["summary"].items()]
    checks_rows = [
        [check["id"], check["type"], check["severity"], check["status"]]
        for check in result["checks"]
    ]
    detail_sections: list[str] = []
    for check in result["checks"]:
        if not check.get("details") and check["status"] == "passed":
            continue
        detail_sections.append(
            f"<section><h3>{html.escape(check['id'])}</h3>"
            f"<p><strong>{html.escape(check['status'].upper())}</strong> · {html.escape(check['type'])}</p>"
            f"<details><summary>Metrics</summary><pre>{html.escape(json.dumps(check['metrics'], indent=2, ensure_ascii=False))}</pre></details>"
            + (
                f"<details open><summary>Evidence sample</summary><pre>{html.escape(json.dumps(check['details'], indent=2, ensure_ascii=False))}</pre></details>"
                if check.get("details")
                else ""
            )
            + "</section>"
        )
    status_class = "passed" if result["status"] == "passed" else "failed"
    return f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Reconciliation evidence — {html.escape(result['reconciliation'])}</title>
<style>
body{{font-family:system-ui,-apple-system,Segoe UI,sans-serif;max-width:1180px;margin:40px auto;padding:0 24px;line-height:1.45;color:#111}}
h1,h2,h3{{line-height:1.15}} .status{{font-weight:700;padding:6px 10px;border:1px solid currentColor;display:inline-block;border-radius:6px}}
.passed{{color:#176b2c}} .failed{{color:#9b1c1c}} table{{border-collapse:collapse;width:100%;margin:12px 0 28px}}th,td{{border:1px solid #ccc;padding:7px 9px;text-align:left;vertical-align:top}}th{{font-weight:650}}pre{{white-space:pre-wrap;overflow-wrap:anywhere;background:#f6f6f6;padding:12px;border-radius:6px}}section{{border-top:1px solid #ddd;padding-top:12px;margin-top:20px}}code{{background:#f4f4f4;padding:2px 4px;border-radius:3px}}
</style>
</head>
<body>
<h1>Reconciliation evidence</h1>
<h2>{html.escape(result['reconciliation'])}</h2>
<p class="status {status_class}">{html.escape(result['status'].upper())}</p>
<p>Run <code>{html.escape(result.get('run', {}).get('id', ''))}</code> · generated {html.escape(result['generated_at'])}</p>
<h2>Summary</h2>
{_table_html(['Metric', 'Value'], summary_rows)}
<h2>Checks</h2>
{_table_html(['Check', 'Type', 'Severity', 'Status'], checks_rows)}
<h2>Details</h2>
{''.join(detail_sections) if detail_sections else '<p>No discrepancy details.</p>'}
<h2>Provenance</h2>
<pre>{html.escape(json.dumps({'inputs': result.get('inputs', {}), 'configuration_sha256': result.get('configuration_sha256'), 'engine_version': result.get('engine_version'), 'schema_version': result.get('schema_version'), 'privacy': result.get('privacy', {})}, indent=2, ensure_ascii=False))}</pre>
</body></html>"""


def write_html(result: dict[str, Any], path: str | Path) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(render_html(result), encoding="utf-8")


def _flatten_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return _json_text(value)
    return value


def _detail_categories(result: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    categories: dict[str, list[dict[str, Any]]] = {
        "missing": [],
        "unexpected": [],
        "field_mismatches": [],
        "aggregate_mismatches": [],
        "exceptions": [],
    }
    for check in result.get("checks", []):
        for detail in check.get("details", []):
            row = {
                "check_id": check["id"],
                "check_type": check["type"],
                "severity": check["severity"],
                **{name: _flatten_value(value) for name, value in detail.items()},
            }
            difference = detail.get("difference")
            if difference == "missing_in_target":
                categories["missing"].append(row)
            elif difference == "unexpected_in_target":
                categories["unexpected"].append(row)
            elif check["type"] == "field_match":
                categories["field_mismatches"].append(row)
            elif check["type"] == "aggregate_match":
                categories["aggregate_mismatches"].append(row)
    return categories


def _write_rows_csv(path: Path, rows: list[dict[str, Any]], default_headers: list[str]) -> None:
    headers = list(default_headers)
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _append_sheet(sheet: Any, headers: list[str], rows: list[dict[str, Any]]) -> None:
    all_headers = list(headers)
    for row in rows:
        for key in row:
            if key not in all_headers:
                all_headers.append(key)
    sheet.append(all_headers)
    for row in rows:
        sheet.append([_flatten_value(row.get(header)) for header in all_headers])
    sheet.freeze_panes = "A2"
    if sheet.max_row >= 1:
        sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = cell.font.copy(bold=True)
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 50)


def write_xlsx(result: dict[str, Any], path: str | Path) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise DataError(
            "XLSX evidence requires openpyxl. Install with: pip install 'reconciliation-as-code[excel]'"
        ) from exc

    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_rows = [
        {"metric": "reconciliation", "value": result["reconciliation"]},
        {"metric": "status", "value": result["status"]},
        {"metric": "run_id", "value": result.get("run", {}).get("id")},
        {"metric": "generated_at", "value": result["generated_at"]},
    ]
    summary_rows.extend({"metric": key, "value": value} for key, value in result["summary"].items())
    _append_sheet(summary_sheet, ["metric", "value"], summary_rows)

    checks_sheet = workbook.create_sheet("Checks")
    checks_rows = [
        {
            "check_id": check["id"],
            "type": check["type"],
            "severity": check["severity"],
            "status": check["status"],
            "metrics": _json_text(check.get("metrics", {})),
            "details_truncated": check.get("details_truncated", False),
        }
        for check in result["checks"]
    ]
    _append_sheet(checks_sheet, ["check_id", "type", "severity", "status", "metrics"], checks_rows)

    categories = _detail_categories(result)
    for sheet_name, category, headers in (
        ("Missing", "missing", ["check_id", "key", "difference"]),
        ("Unexpected", "unexpected", ["check_id", "key", "difference"]),
        ("Field Mismatches", "field_mismatches", ["check_id", "key", "source", "target"]),
        ("Aggregate Mismatches", "aggregate_mismatches", ["check_id", "group", "source_value", "target_value"]),
        ("Exceptions", "exceptions", ["check_id", "key", "reason", "reference"]),
    ):
        sheet = workbook.create_sheet(sheet_name)
        _append_sheet(sheet, headers, categories[category])

    totals_sheet = workbook.create_sheet("Totals")
    totals_rows = []
    for check in result["checks"]:
        if check["type"] in {"control_total", "aggregate_match", "row_count"}:
            totals_rows.append(
                {
                    "check_id": check["id"],
                    "type": check["type"],
                    "status": check["status"],
                    "metrics": _json_text(check.get("metrics", {})),
                }
            )
    _append_sheet(totals_sheet, ["check_id", "type", "status", "metrics"], totals_rows)
    workbook.save(output)


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_bundle(result: dict[str, Any], path: str | Path) -> dict[str, Any]:
    bundle = Path(path)
    bundle.mkdir(parents=True, exist_ok=True)
    details_dir = bundle / "details"
    details_dir.mkdir(parents=True, exist_ok=True)

    write_json(result, bundle / "evidence.json")
    write_markdown(result, bundle / "evidence.md")
    write_html(result, bundle / "evidence.html")
    write_xlsx(result, bundle / "evidence.xlsx")

    categories = _detail_categories(result)
    _write_rows_csv(details_dir / "missing.csv", categories["missing"], ["check_id", "key", "difference"])
    _write_rows_csv(details_dir / "unexpected.csv", categories["unexpected"], ["check_id", "key", "difference"])
    _write_rows_csv(
        details_dir / "field-mismatches.csv",
        categories["field_mismatches"],
        ["check_id", "key", "source", "target"],
    )
    _write_rows_csv(
        details_dir / "aggregate-mismatches.csv",
        categories["aggregate_mismatches"],
        ["check_id", "group", "source_value", "target_value"],
    )
    _write_rows_csv(
        details_dir / "exceptions.csv",
        categories["exceptions"],
        ["check_id", "key", "reason", "reference"],
    )

    generated_files: dict[str, dict[str, Any]] = {}
    for file_path in sorted(bundle.rglob("*")):
        if not file_path.is_file() or file_path.name == "manifest.json":
            continue
        relative = file_path.relative_to(bundle).as_posix()
        generated_files[relative] = {
            "sha256": _file_sha256(file_path),
            "size_bytes": file_path.stat().st_size,
        }

    manifest = {
        "bundle_version": 1,
        "run_id": result.get("run", {}).get("id"),
        "reconciliation": result["reconciliation"],
        "status": result["status"],
        "generated_at": result["generated_at"],
        "engine_version": result.get("engine_version"),
        "schema_version": result.get("schema_version"),
        "configuration_sha256": result.get("configuration_sha256"),
        "inputs": result.get("inputs", {}),
        "privacy": result.get("privacy", {}),
        "files": generated_files,
    }
    write_json(manifest, bundle / "manifest.json")
    return manifest
