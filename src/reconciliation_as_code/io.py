from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from .errors import DataError


def _validate_headers(headers: list[str], path: Path, label: str, *, case_insensitive: bool = False) -> None:
    if not headers:
        raise DataError(f"{label} has no header: {path}")
    if any(not header.strip() for header in headers):
        raise DataError(f"{label} header contains an empty column name: {path}")
    seen: set[str] = set()
    for header in headers:
        key = header.casefold() if case_insensitive else header
        if key in seen:
            raise DataError(f"{label} header contains a duplicate column name {header!r}: {path}")
        seen.add(key)


def validate_csv_header(path: Path, delimiter: str = ",", *, case_insensitive: bool = False) -> None:
    """Check original column names before a backend can rename duplicates."""
    if not isinstance(delimiter, str) or len(delimiter) != 1:
        raise DataError("CSV delimiter must be exactly one character.")
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            headers = next(csv.reader(handle, delimiter=delimiter, strict=True), [])
            _validate_headers(headers, path, "CSV", case_insensitive=case_insensitive)
    except UnicodeDecodeError as exc:
        raise DataError(f"CSV must be UTF-8 encoded: {path}") from exc
    except csv.Error as exc:
        raise DataError(f"Invalid CSV header: {path}: {exc}") from exc


def _load_csv(path: Path, delimiter: str = ",") -> list[dict[str, Any]]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle, delimiter=delimiter, strict=True)
            _validate_headers(reader.fieldnames or [], path, "CSV")
            rows = []
            for row in reader:
                if None in row or any(value is None for value in row.values()):
                    raise DataError(
                        f"CSV row at line {reader.line_num} does not match the header width: {path}"
                    )
                rows.append(dict(row))
            return rows
    except UnicodeDecodeError as exc:
        raise DataError(f"CSV must be UTF-8 encoded: {path}") from exc
    except csv.Error as exc:
        raise DataError(f"Invalid CSV at line {reader.line_num}: {path}: {exc}") from exc


def _load_excel(path: Path, sheet: str | None = None) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise DataError(
            "Excel input requires the optional dependency. Install with: pip install 'reconciliation-as-code[excel]'"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet:
            if sheet not in workbook.sheetnames:
                raise DataError(f"Sheet {sheet!r} not found in {path}. Available: {workbook.sheetnames}")
            worksheet = workbook[sheet]
        else:
            worksheet = workbook[workbook.sheetnames[0]]

        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
        _validate_headers(headers, path, "Excel")
        return [dict(zip(headers, row, strict=False)) for row in rows]
    finally:
        workbook.close()


def load_table(endpoint: dict[str, Any], base_dir: Path) -> tuple[Path, list[dict[str, Any]]]:
    path = Path(endpoint["file"])
    if not path.is_absolute():
        path = base_dir / path
    path = path.resolve()
    if not path.exists():
        raise DataError(f"Input file not found: {path}")

    file_format = str(endpoint.get("format") or path.suffix.lstrip(".")).lower()
    if file_format == "csv":
        delimiter = endpoint.get("delimiter", ",")
        if not isinstance(delimiter, str) or len(delimiter) != 1:
            raise DataError("CSV delimiter must be exactly one character.")
        rows = _load_csv(path, delimiter)
    elif file_format in {"xlsx", "xlsm"}:
        rows = _load_excel(path, endpoint.get("sheet"))
    else:
        raise DataError(f"Unsupported input format {file_format!r} for {path}")

    return path, rows
