from __future__ import annotations

import csv
import hashlib
import json
import tempfile
import unittest
from pathlib import Path

import yaml
from openpyxl import load_workbook

from reconciliation_as_code.cli import main as cli_main
from reconciliation_as_code.engine import run_reconciliation
from reconciliation_as_code.report import prepare_evidence, write_bundle
from reconciliation_as_code.spec import validate_spec


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def privacy_spec() -> dict:
    return {
        "version": 1,
        "reconciliation": {"name": "privacy evidence"},
        "source": {"file": "source.csv", "key": "CUSTOMER_ID"},
        "target": {"file": "target.csv", "key": "CUSTOMER_ID"},
        "checks": [
            {"id": "coverage", "type": "record_coverage"},
            {
                "id": "email",
                "type": "field_match",
                "source": "EMAIL",
                "target": "EMAIL",
            },
            {
                "id": "balance",
                "type": "control_total",
                "source": "BALANCE",
                "target": "BALANCE",
                "tolerance": 0,
            },
        ],
        "evidence": {
            "detail_limit": 100,
            "sensitive_fields": ["EMAIL"],
            "sensitive_value_mode": "hash",
            "key_mode": "hash",
        },
    }


class EvidenceBundleTests(unittest.TestCase):
    def test_prepare_evidence_redacts_values_and_hashes_keys(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            write_csv(
                base / "source.csv",
                [{"CUSTOMER_ID": "4711", "EMAIL": "alice@example.com", "BALANCE": "100"}],
            )
            write_csv(
                base / "target.csv",
                [{"CUSTOMER_ID": "4711", "EMAIL": "wrong@example.com", "BALANCE": "100"}],
            )
            spec = privacy_spec()
            raw = run_reconciliation(spec, base_dir=base)
            prepared = prepare_evidence(raw, spec)
            serialized = json.dumps(prepared)
            self.assertNotIn("alice@example.com", serialized)
            self.assertNotIn("wrong@example.com", serialized)
            self.assertNotIn('"key": "4711"', serialized)
            email_check = next(item for item in prepared["checks"] if item["id"] == "email")
            self.assertTrue(str(email_check["details"][0]["source"]).startswith("sha256:"))
            self.assertTrue(str(email_check["details"][0]["key"]).startswith("sha256:"))

    def test_one_cli_run_creates_complete_offline_bundle(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            write_csv(
                base / "source.csv",
                [
                    {"CUSTOMER_ID": "4711", "EMAIL": "alice@example.com", "BALANCE": "100"},
                    {"CUSTOMER_ID": "5000", "EMAIL": "bob@example.com", "BALANCE": "50"},
                ],
            )
            write_csv(
                base / "target.csv",
                [
                    {"CUSTOMER_ID": "4711", "EMAIL": "wrong@example.com", "BALANCE": "95"},
                    {"CUSTOMER_ID": "6000", "EMAIL": "eve@example.com", "BALANCE": "55"},
                ],
            )
            spec = privacy_spec()
            spec_path = base / "reconciliation.yaml"
            spec_path.write_text(yaml.safe_dump(spec, sort_keys=False), encoding="utf-8")
            bundle = base / "bundle"
            evidence = base / "standalone.json"
            report = base / "standalone.md"

            code = cli_main(
                [
                    "run",
                    str(spec_path),
                    "--bundle",
                    str(bundle),
                    "--evidence",
                    str(evidence),
                    "--report",
                    str(report),
                    "--no-fail-on-diff",
                ]
            )
            self.assertEqual(0, code)

            expected_files = {
                "evidence.json",
                "evidence.md",
                "evidence.html",
                "evidence.xlsx",
                "manifest.json",
                "details/missing.csv",
                "details/unexpected.csv",
                "details/field-mismatches.csv",
                "details/aggregate-mismatches.csv",
                "details/exceptions.csv",
            }
            actual_files = {
                path.relative_to(bundle).as_posix() for path in bundle.rglob("*") if path.is_file()
            }
            self.assertTrue(expected_files <= actual_files)

            for text_path in [
                bundle / "evidence.json",
                bundle / "evidence.md",
                bundle / "evidence.html",
                bundle / "manifest.json",
                bundle / "details" / "field-mismatches.csv",
            ]:
                text = text_path.read_text(encoding="utf-8")
                for raw_value in (
                    "alice@example.com",
                    "wrong@example.com",
                    "bob@example.com",
                    "eve@example.com",
                ):
                    self.assertNotIn(raw_value, text)

            html_text = (bundle / "evidence.html").read_text(encoding="utf-8")
            self.assertNotIn('src="http', html_text)
            self.assertNotIn('href="http', html_text)
            self.assertNotIn("<script", html_text.lower())

            workbook = load_workbook(bundle / "evidence.xlsx", read_only=True)
            required_sheets = {
                "Summary",
                "Checks",
                "Missing",
                "Unexpected",
                "Field Mismatches",
                "Totals",
                "Exceptions",
            }
            self.assertTrue(required_sheets <= set(workbook.sheetnames))
            workbook_values = " ".join(
                str(cell.value or "")
                for sheet in workbook.worksheets
                for row in sheet.iter_rows()
                for cell in row
            )
            self.assertNotIn("alice@example.com", workbook_values)
            self.assertNotIn("wrong@example.com", workbook_values)
            workbook.close()

            manifest = json.loads((bundle / "manifest.json").read_text(encoding="utf-8"))
            self.assertEqual(1, manifest["bundle_version"])
            self.assertIn("specification", manifest["inputs"])
            self.assertIn("evidence.json", manifest["files"])
            evidence_bytes = (bundle / "evidence.json").read_bytes()
            self.assertEqual(
                hashlib.sha256(evidence_bytes).hexdigest(),
                manifest["files"]["evidence.json"]["sha256"],
            )

    def test_bundle_can_be_written_from_prepared_evidence(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            write_csv(base / "source.csv", [{"CUSTOMER_ID": "1", "EMAIL": "a@x", "BALANCE": "1"}])
            write_csv(base / "target.csv", [{"CUSTOMER_ID": "1", "EMAIL": "b@x", "BALANCE": "1"}])
            spec = privacy_spec()
            result = prepare_evidence(run_reconciliation(spec, base_dir=base), spec)
            manifest = write_bundle(result, base / "bundle")
            self.assertEqual(result["run"]["id"], manifest["run_id"])
            self.assertGreaterEqual(len(manifest["files"]), 9)

    def test_invalid_privacy_mode_is_rejected(self) -> None:
        spec = privacy_spec()
        spec["evidence"]["sensitive_value_mode"] = "encrypt-maybe"
        with self.assertRaisesRegex(Exception, "sensitive_value_mode"):
            validate_spec(spec)


if __name__ == "__main__":
    unittest.main()
